
import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
from pathlib import Path
from fpdf import FPDF

# === CONFIGURATION ===
PASSWORD = "admin123"
file_path = "Sommaire_Projet_Reserves.xlsx"
columns = ["N° Réserve", "Description", "Date de création", "Date de levée", "Statut"]

# === AUTHENTIFICATION ===
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("🔐 Authentification requise")
    password = st.text_input("Entrez le mot de passe", type="password")
    if password == PASSWORD:
        st.session_state.authenticated = True
        st.rerun()
    elif password:
        st.error("Mot de passe incorrect")
    st.stop()

# === CHARGER LES DONNÉES ===
wb = openpyxl.load_workbook(file_path)
ws = wb["Réserves"]

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(row):
        data.append(row)

df = pd.DataFrame(data, columns=columns)

st.title("🛠️ Interface de gestion des réserves")
st.markdown("Visualisez, filtrez, modifiez, exportez les réserves du projet.")

# === FILTRES ===
col1, col2 = st.columns(2)

with col1:
    filtre_statut = st.selectbox("Filtrer par statut", options=["Tous"] + sorted(df["Statut"].dropna().unique().tolist()))
with col2:
    filtre_date = st.date_input("Filtrer par date de création après", value=None)

df_filtered = df.copy()

if filtre_statut != "Tous":
    df_filtered = df_filtered[df_filtered["Statut"] == filtre_statut]

if filtre_date:
    df_filtered = df_filtered[pd.to_datetime(df_filtered["Date de création"], errors='coerce') >= pd.to_datetime(filtre_date)]

# === EXPORT PDF ===
def export_to_pdf(df):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    pdf.set_fill_color(200, 220, 255)

    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "Liste des Réserves", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Arial", "B", 10)
    for col in columns:
        pdf.cell(38, 8, col, 1, 0, 'C', fill=True)
    pdf.ln()

    pdf.set_font("Arial", "", 10)
    for _, row in df.iterrows():
        for item in row:
            pdf.cell(38, 8, str(item)[:35], 1)
        pdf.ln()

    output_pdf = "export_reserves.pdf"
    pdf.output(output_pdf)
    return output_pdf

if st.button("📄 Exporter en PDF"):
    pdf_path = export_to_pdf(df_filtered)
    with open(pdf_path, "rb") as f:
        st.download_button("Télécharger le PDF", data=f, file_name="reserves.pdf", mime="application/pdf")

# === TABLEAU ÉDITABLE ===
st.subheader("🔧 Modifier les réserves existantes")
edited_df = st.experimental_data_editor(df_filtered, num_rows="dynamic")

# === AJOUTER UNE NOUVELLE RÉSERVE ===
st.subheader("➕ Ajouter une nouvelle réserve")

with st.form("ajout_reserve"):
    new_id = st.text_input("N° Réserve")
    new_desc = st.text_area("Description")
    new_date_crea = st.date_input("Date de création", value=datetime.today())
    new_date_lev = st.date_input("Date de levée", value=None)
    new_statut = st.selectbox("Statut", ["", "en cours", "levée", "N/A"])
    submit = st.form_submit_button("Ajouter")

    if submit:
        if new_id:
            new_row = [new_id, new_desc, new_date_crea.strftime("%Y-%m-%d"), new_date_lev.strftime("%Y-%m-%d") if new_date_lev else "", new_statut]
            ws.append(new_row)
            wb.save(file_path)
            st.success(f"Réserve {new_id} ajoutée avec succès !")

# === SAUVEGARDE DES MODIFICATIONS DU TABLEAU ===
if st.button("💾 Enregistrer les modifications affichées"):
    for i, row in edited_df.iterrows():
        index_in_sheet = df[df["N° Réserve"] == row["N° Réserve"]].index
        if not index_in_sheet.empty:
            excel_row = index_in_sheet[0] + 2
            for j, col in enumerate(columns):
                ws.cell(row=excel_row, column=j+1).value = row[col] if row[col] != "" else None
    wb.save(file_path)
    st.success("Modifications enregistrées avec succès.")
